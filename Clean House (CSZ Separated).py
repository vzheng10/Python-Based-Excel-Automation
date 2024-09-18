   #import openpyxl and import existing worbook
from openpyxl import Workbook, load_workbook
    #import row|col index function from openpyxl
from openpyxl.utils import get_column_letter
    #import to style and align cells 
from openpyxl.styles import Font, Alignment
    #import to extract path name -> print code
from pathlib import Path
 #import tkinter -> provides a GUI for file selection
import tkinter as tk
from tkinter import filedialog
    
#1. Clean HOUSE (City, State, Zip SEPARATED) Data File *************************

print(f"Select Data File")
#create a root window and hide it
root = tk.Tk()
root.withdraw() #hides the root window

#open a file dialog and prompt user to select Excel file
file_path = filedialog.askopenfilename(
    title = "Select a kill file",
    filetypes = [("Excel files", "*.xlsx *xls")]
)

if file_path:
    try:
            #workbook name -> printcode
        wb_path = Path(file_path)
            #load an existing workbook (wb)
        wb = load_workbook(wb_path) #&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
            #accessing worksheets (ws) within a workbook
        ws = wb.active
            #extract workbook name from path
        print_code = wb_path.stem

            #create PrintCode, PrintOrder and QR column
        ws['H1'].value = "PrintCode"    # -> #8
        ws['I1'].value = "PrintOrder"   # -> #9
        ws['J1'].value = "QR"           # -> #10

            #Col H = 8, Col J = 10
        for col in range (8,11):
            cell =  ws[get_column_letter(col) + '1']
            cell.font = Font(bold = True)
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            #create printcode and QR
        for row in range (2, ws.max_row + 1):
            #create PrintCode in Col H
            ws[f'H{row}'] = print_code

            #create PrintOrder in Col I
            ws[f'I{row}'] = row - 1

            #concatenate info
            formula = f'=TRIM(A{row})&";;"&IF(C{row}="","",B{row})&";"&IF(C{row}="",B{row},C{row})&";"&D{row}&", "&E{row}&"  "&F{row}&";"&H{row}'
            ws[f'J{row}'].value = formula

        #make a new worksheet called Count
        ws_count = wb.create_sheet("Count")

        ws_count[f'A1'].value = print_code
        ws_count[f'B1'].value = "=COUNTIF(Sheet1!H:H,Count!A1)"

        ws_count[f'A3'].value = "Total"
        ws_count[f'B3'].value = f"=SUM({'B1'})"

        wb.save(wb_path) #&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
        
        print(f"House Data Clean Done")

    except Exception as e:
        print(f"An error occured: {e}")

else:
    print("No file was selected.")

root.destroy()


